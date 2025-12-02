from pathlib import Path
from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parents[1]
INPUT = ROOT / 'Documentacion' / 'lfpiorpi2025.xlsx'
OUT = ROOT / 'Documentacion' / 'avisos2025_summaries' / 'lfpiorpi2025.md'

def sample_sheet(ws, max_rows=5):
    rows = []
    for i, row in enumerate(ws.iter_rows(values_only=True)):
        # collect up to header + max_rows data rows
        rows.append([cell if cell is not None else '' for cell in row])
        if i >= max_rows:
            break
    return rows

def to_md_table(rows):
    if not rows:
        return '_(hoja vacía)_\n\n'
    # use first row as header
    header = rows[0]
    body = rows[1:]
    md = '| ' + ' | '.join(str(h) for h in header) + ' |\n'
    md += '| ' + ' | '.join('---' for _ in header) + ' |\n'
    for r in body:
        md += '| ' + ' | '.join(str(c) for c in r) + ' |\n'
    md += '\n'
    return md

def main():
    if not INPUT.exists():
        print(f'ERROR: no existe {INPUT}')
        return

    wb = load_workbook(INPUT, read_only=True, data_only=True)
    parts = []
    parts.append('# lfpiorpi2025.xlsx\n')
    parts.append(f'- Ruta: `Documentacion/lfpiorpi2025.xlsx`\n')
    parts.append('- Tipo: .xlsx\n\n')
    parts.append('## Hojas y muestra de contenido\n')

    for sheetname in wb.sheetnames:
        ws = wb[sheetname]
        parts.append(f'### {sheetname}\n\n')
        rows = sample_sheet(ws, max_rows=6)
        # if only one row, still show it
        parts.append(to_md_table(rows))

    OUT.parent.mkdir(parents=True, exist_ok=True)
    OUT.write_text('\n'.join(parts), encoding='utf-8')
    print(f'Wrote summary to {OUT}')

if __name__ == '__main__':
    main()
